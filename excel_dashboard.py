"""
Excel Dashboard Generator for ETE IT PMO Resource Planning.

Aligned with Jim K's Resource Model Template:
- Per-project role × SDLC phase hours matrices
- ALL values are FORMULAS referencing source sheets (not hardcoded)
- Only hardcoded value per project = the Project ID (lookup key)
- Everything else derives from Project Portfolio, RM_Assumptions, Team Roster

Source sheets (Project Portfolio, Team Roster, RM_Assumptions) are NEVER modified.
"""

import os
from datetime import date, timedelta
from collections import defaultdict, Counter
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

from excel_connector import ExcelConnector, SDLC_PHASES
from capacity_engine import CapacityEngine


# ── Colors ──────────────────────────────────────────────────────────
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
GOLD_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
LIGHT_BLUE_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
LIGHT_GREEN_FILL = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
LIGHT_GRAY_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

PRIORITY_FILLS = {
    "Highest": PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
    "High": PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid"),
    "Medium": PatternFill(start_color="6BCB77", end_color="6BCB77", fill_type="solid"),
    "Low": PatternFill(start_color="4D96FF", end_color="4D96FF", fill_type="solid"),
}

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BOLD_FONT = Font(bold=True, size=11)
NORMAL_FONT = Font(size=10)
SMALL_FONT = Font(size=9, color="666666")
TITLE_FONT = Font(bold=True, size=14, color="4472C4")
SECTION_FONT = Font(bold=True, size=12, color="4472C4")
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

# ── Role configuration ──────────────────────────────────────────────
# Each tuple: (role_key, display_label, portfolio_col, assumptions_row, roster_role_name)
# portfolio_col = column letter in Project Portfolio for that role's allocation %
# assumptions_row = row in RM_Assumptions for that role's phase effort profile
# roster_role_name = role label in Team Roster column B (for AVERAGEIFS)
ROLE_CONFIG = [
    ("pm",             "Project Manager",          "Z",  21, "Project Manager"),
    ("dba",            "DBA",                      "Y",  22, "DBA"),
    ("ba",             "IT Business Analyst/SME",  "T",  23, "Business Analyst"),
    ("functional",     "IT Functional Analyst",    "U",  24, "Functional"),
    ("technical",      "IT Tech Analyst/Developer","V",  25, "Technical"),
    ("developer",      "Developer",                "W",  26, "Developer"),
    ("infrastructure", "Infrastructure",           "X",  27, "Infrastructure"),
    ("wms",            "WMS Consultant",            "AA", 28, "WMS Consultant"),
]

PHASE_LABELS = {
    "discovery": "Discovery", "planning": "Planning", "design": "Design",
    "build": "Build", "test": "Test", "deploy": "Deploy/Hypercare",
}

# RM_Assumptions phase weight columns: B=Discovery, C=Planning, D=Design, E=Build, F=Test, G=Deploy
PHASE_ASSUMPTION_COLS = {"discovery": "B", "planning": "C", "design": "D",
                          "build": "E", "test": "F", "deploy": "G"}


def _util_fill(pct):
    if pct >= 1.0:
        return RED_FILL
    elif pct >= 0.80:
        return YELLOW_FILL
    return GREEN_FILL


def _write_header_row(ws, row, headers, col_start=1):
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=col_start + i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def _style_data_cell(cell, center=True):
    cell.border = THIN_BORDER
    cell.font = NORMAL_FONT
    if center:
        cell.alignment = Alignment(horizontal="center")


def _safe_delete(wb, name):
    if name in wb.sheetnames:
        del wb[name]


class DashboardGenerator:

    def __init__(self, connector=None):
        self.connector = connector or ExcelConnector()
        self.engine = CapacityEngine(self.connector)
        self._data = self.engine._load()

    def generate_all(self, output_path=None):
        wb = openpyxl.load_workbook(self.connector.workbook_path)

        # Remove old computed sheets
        for name in ["Dashboard", "Capacity Heatmap", "Resource Scheduler",
                      "RM_Demand", "RM_SupplyVsDemand", "RM_WeeklyHeatmap",
                      "RM_Gantt", "Project Timeline", "Resource Model",
                      "Capacity Summary", "Role Capacity Planner"]:
            _safe_delete(wb, name)

        # Generate formula-driven sheets
        self._write_resource_model(wb)
        self._write_capacity_summary(wb)
        self._write_role_capacity_planner(wb)
        self._write_capacity_heatmap(wb)
        self._write_gantt(wb)

        # Reorder
        desired_order = [
            "Resource Model", "Capacity Summary", "Role Capacity Planner",
            "Project Portfolio", "Team Roster", "RM_Assumptions",
            "Capacity Heatmap", "RM_Gantt",
            "RM_Guide", "AI_Agent_Spec",
        ]
        for i, name in enumerate(desired_order):
            if name in wb.sheetnames:
                idx = wb.sheetnames.index(name)
                wb.move_sheet(name, offset=i - idx)

        if output_path is None:
            base = Path(self.connector.workbook_path)
            output_path = str(base.parent / f"{base.stem}_Dashboard{base.suffix}")

        wb.save(output_path)
        print(f"Dashboard saved: {output_path}")
        return output_path

    # ================================================================
    # RESOURCE MODEL — Formula-driven, Jim K-style
    # ================================================================
    def _write_resource_model(self, wb):
        ws = wb.create_sheet("Resource Model")

        active = self._data["active_portfolio"]

        # ────────────────────────────────────────────────────────────
        # BLOCK 1: Configuration (formulas referencing RM_Assumptions)
        # ────────────────────────────────────────────────────────────
        ws["A1"] = "Resource Model"
        ws["A1"].font = TITLE_FONT
        ws["A2"] = "All values are formulas — change source data and this sheet recalculates"
        ws["A2"].font = SMALL_FONT

        # ── Time Allocation ──
        row = 4
        ws.cell(row=row, column=1, value="Time Allocation").font = SECTION_FONT
        row += 1
        for label, formula in [
            ("Base Hours/Week",       "=RM_Assumptions!B6"),
            ("Admin %",              "=RM_Assumptions!B7"),
            ("Support/Break-Fix %",  "=RM_Assumptions!B8"),
            ("Project %",           "=RM_Assumptions!B9"),
            ("Project Hrs/Person/Wk","=RM_Assumptions!B10"),
            ("Max Projects/Person",  "=RM_Assumptions!B12"),
        ]:
            ws.cell(row=row, column=1, value=label).font = NORMAL_FONT
            ws.cell(row=row, column=1).border = THIN_BORDER
            cell = ws.cell(row=row, column=2)
            cell.value = formula
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
            if "%" in label:
                cell.number_format = '0%'
            row += 1

        # ── SDLC Phase Weights ──
        row += 1
        ws.cell(row=row, column=1, value="SDLC Phase Weights").font = SECTION_FONT
        row += 1
        # Headers
        ws.cell(row=row, column=1, value="Phase").font = BOLD_FONT
        ws.cell(row=row, column=1).fill = LIGHT_BLUE_FILL
        ws.cell(row=row, column=1).border = THIN_BORDER
        for i, phase in enumerate(SDLC_PHASES):
            cell = ws.cell(row=row, column=2 + i, value=PHASE_LABELS.get(phase, phase))
            cell.font = BOLD_FONT
            cell.fill = LIGHT_BLUE_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER
        ws.cell(row=row, column=8, value="Total").font = BOLD_FONT
        ws.cell(row=row, column=8).fill = LIGHT_BLUE_FILL
        ws.cell(row=row, column=8).border = THIN_BORDER

        # Weight values (formulas)
        phase_weight_row = row + 1
        ws.cell(row=phase_weight_row, column=1, value="Weight").font = NORMAL_FONT
        ws.cell(row=phase_weight_row, column=1).border = THIN_BORDER
        for i, phase in enumerate(SDLC_PHASES):
            acol = PHASE_ASSUMPTION_COLS[phase]
            cell = ws.cell(row=phase_weight_row, column=2 + i)
            cell.value = f"=RM_Assumptions!{acol}16"
            cell.number_format = '0%'
            _style_data_cell(cell)
        ws.cell(row=phase_weight_row, column=8).value = f"=SUM(B{phase_weight_row}:G{phase_weight_row})"
        ws.cell(row=phase_weight_row, column=8).number_format = '0%'
        ws.cell(row=phase_weight_row, column=8).font = BOLD_FONT
        ws.cell(row=phase_weight_row, column=8).border = THIN_BORDER

        row = phase_weight_row + 2

        # ── Role Effort Profiles ──
        ws.cell(row=row, column=1, value="Role Effort by SDLC Phase").font = SECTION_FONT
        row += 1
        ws.cell(row=row, column=1, value="Role").font = BOLD_FONT
        ws.cell(row=row, column=1).fill = LIGHT_BLUE_FILL
        ws.cell(row=row, column=1).border = THIN_BORDER
        for i, phase in enumerate(SDLC_PHASES):
            cell = ws.cell(row=row, column=2 + i, value=PHASE_LABELS.get(phase, phase))
            cell.font = BOLD_FONT
            cell.fill = LIGHT_BLUE_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER
        ws.cell(row=row, column=8, value="Total").font = BOLD_FONT
        ws.cell(row=row, column=8).fill = LIGHT_BLUE_FILL
        ws.cell(row=row, column=8).border = THIN_BORDER

        row += 1
        role_effort_start_row = row
        for role_key, label, port_col, assump_row, roster_name in ROLE_CONFIG:
            ws.cell(row=row, column=1, value=label).font = NORMAL_FONT
            ws.cell(row=row, column=1).border = THIN_BORDER
            for i, phase in enumerate(SDLC_PHASES):
                acol = PHASE_ASSUMPTION_COLS[phase]
                cell = ws.cell(row=row, column=2 + i)
                cell.value = f"=RM_Assumptions!{acol}{assump_row}"
                cell.number_format = '0%'
                _style_data_cell(cell)
            ws.cell(row=row, column=8).value = f"=SUM(B{row}:G{row})"
            ws.cell(row=row, column=8).number_format = '0%'
            ws.cell(row=row, column=8).font = BOLD_FONT
            ws.cell(row=row, column=8).border = THIN_BORDER
            row += 1

        row += 1

        # ── Per-Person Capacity (AVERAGEIFS from Team Roster) ──
        ws.cell(row=row, column=1, value="Per-Person Capacity").font = SECTION_FONT
        row += 1
        for col, label in [(1, "Role"), (2, "Headcount"), (3, "Team Supply\n(hrs/wk)"),
                            (4, "Per-Person\n(hrs/wk)")]:
            cell = ws.cell(row=row, column=col, value=label)
            cell.font = BOLD_FONT
            cell.fill = LIGHT_BLUE_FILL
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = THIN_BORDER

        row += 1
        # Store the row where each role's per-person capacity is, for project block references
        capacity_rows = {}  # role_key → row number
        for role_key, label, port_col, assump_row, roster_name in ROLE_CONFIG:
            capacity_rows[role_key] = row
            ws.cell(row=row, column=1, value=label).font = NORMAL_FONT
            ws.cell(row=row, column=1).border = THIN_BORDER

            # Headcount = COUNTIF on Team Roster role column
            cell = ws.cell(row=row, column=2)
            cell.value = f'=COUNTIF(\'Team Roster\'!$B$4:$B$26,"{roster_name}")'
            _style_data_cell(cell)

            # Team supply = SUMIFS on Team Roster capacity
            cell = ws.cell(row=row, column=3)
            cell.value = f'=SUMIFS(\'Team Roster\'!$K$4:$K$26,\'Team Roster\'!$B$4:$B$26,"{roster_name}")'
            cell.number_format = '0.0'
            _style_data_cell(cell)

            # Per-person = AVERAGEIFS
            cell = ws.cell(row=row, column=4)
            cell.value = f'=IFERROR(AVERAGEIFS(\'Team Roster\'!$K$4:$K$26,\'Team Roster\'!$B$4:$B$26,"{roster_name}"),0)'
            cell.number_format = '0.0'
            _style_data_cell(cell)
            cell.fill = GOLD_FILL

            row += 1

        # Totals
        cap_start = capacity_rows[ROLE_CONFIG[0][0]]
        cap_end = row - 1
        ws.cell(row=row, column=1, value="TOTAL").font = BOLD_FONT
        ws.cell(row=row, column=1).fill = LIGHT_BLUE_FILL
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2).value = f"=SUM(B{cap_start}:B{cap_end})"
        ws.cell(row=row, column=2).font = BOLD_FONT
        ws.cell(row=row, column=2).fill = LIGHT_BLUE_FILL
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=3).value = f"=SUM(C{cap_start}:C{cap_end})"
        ws.cell(row=row, column=3).font = BOLD_FONT
        ws.cell(row=row, column=3).fill = LIGHT_BLUE_FILL
        ws.cell(row=row, column=3).border = THIN_BORDER
        ws.cell(row=row, column=3).number_format = '0.0'

        row += 3

        # ────────────────────────────────────────────────────────────
        # BLOCK 2: Per-Project Resource Breakdown (formula-driven)
        # ────────────────────────────────────────────────────────────
        ws.cell(row=row, column=1, value="Per-Project Resource Breakdown").font = TITLE_FONT
        row += 1
        ws.cell(row=row, column=1,
                value="Only the Project ID is entered — all other values are formulas referencing source sheets"
                ).font = SMALL_FONT
        row += 2

        priority_order = {"Highest": 0, "High": 1, "Medium": 2, "Low": 3}
        sorted_projects = sorted(active,
                                  key=lambda p: (priority_order.get(p.priority, 9), p.name))

        for project in sorted_projects:
            active_roles = {k: v for k, v in project.role_allocations.items() if v > 0}
            if not active_roles or project.est_hours <= 0:
                continue

            row = self._write_project_block_formulas(
                ws, row, project.id, active_roles,
                phase_weight_row, capacity_rows
            )

        # Column widths
        ws.column_dimensions["A"].width = 30
        for i in range(6):
            ws.column_dimensions[get_column_letter(2 + i)].width = 14
        ws.column_dimensions["H"].width = 12
        ws.column_dimensions["I"].width = 12
        ws.column_dimensions["J"].width = 14
        ws.column_dimensions["K"].width = 14

    def _write_project_block_formulas(self, ws, row, project_id, active_roles,
                                       phase_weight_row, capacity_rows):
        """
        Write a per-project role × phase matrix using ONLY formulas.
        The project_id is the sole hardcoded value — everything else is
        INDEX/MATCH lookups into Project Portfolio, RM_Assumptions, Team Roster.
        """
        header_row = row

        # ── Row 0: Project header ──
        # Column A: Project ID (the ONLY hardcoded value)
        ws.cell(row=row, column=1, value=project_id).font = Font(bold=True, size=12)
        ws.cell(row=row, column=1).border = THIN_BORDER

        # Column B: Project Name (formula)
        cell = ws.cell(row=row, column=2)
        cell.value = f"=INDEX('Project Portfolio'!$B:$B,MATCH($A{row},'Project Portfolio'!$A:$A,0))"
        cell.font = Font(bold=True, size=12)
        cell.border = THIN_BORDER

        # Column E: Priority (formula)
        cell = ws.cell(row=row, column=5)
        cell.value = f"=INDEX('Project Portfolio'!$H:$H,MATCH($A{header_row},'Project Portfolio'!$A:$A,0))"
        cell.font = BOLD_FONT
        cell.border = THIN_BORDER
        ws.cell(row=row, column=4, value="Priority:").font = SMALL_FONT

        # Column G: Est Hours (formula) — this is the key cell other formulas reference
        cell = ws.cell(row=row, column=7)
        cell.value = f"=INDEX('Project Portfolio'!$R:$R,MATCH($A{header_row},'Project Portfolio'!$A:$A,0))"
        cell.font = Font(bold=True, size=12)
        cell.border = THIN_BORDER
        cell.number_format = '#,##0'
        ws.cell(row=row, column=6, value="Est Hrs:").font = SMALL_FONT

        est_cell = f"$G${header_row}"  # reference for formulas below

        row += 2

        # ── Column headers ──
        headers = ["", "Alloc %"]
        headers += [PHASE_LABELS.get(p, p) for p in SDLC_PHASES]
        headers += ["Total Hrs", "Days"]
        for i, h in enumerate(headers):
            cell = ws.cell(row=row, column=1 + i, value=h)
            cell.font = BOLD_FONT
            cell.fill = LIGHT_BLUE_FILL
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = THIN_BORDER
        col_header_row = row

        row += 1

        # ── Phase Hours row: Est × Phase% ──
        ws.cell(row=row, column=1, value="Phase Hours").font = BOLD_FONT
        ws.cell(row=row, column=1).fill = LIGHT_GRAY_FILL
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2, value="").border = THIN_BORDER  # no alloc for this row

        for i, phase in enumerate(SDLC_PHASES):
            col = 3 + i  # C through H
            cell = ws.cell(row=row, column=col)
            # phase hours = est_hours × phase_weight
            cell.value = f"={est_cell}*{get_column_letter(2+i)}{phase_weight_row}"
            cell.number_format = '0.0'
            cell.fill = LIGHT_GRAY_FILL
            _style_data_cell(cell)

        # Total
        cell = ws.cell(row=row, column=9)
        cell.value = f"=SUM(C{row}:H{row})"
        cell.number_format = '0.0'
        cell.font = BOLD_FONT
        cell.fill = LIGHT_GRAY_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

        phase_hours_row = row
        row += 1

        # ── Role rows ──
        role_rows = []  # track (row_num, role_key) for duration/bottleneck calc

        for role_key, label, port_col, assump_row, roster_name in ROLE_CONFIG:
            if role_key not in active_roles:
                continue

            role_rows.append((row, role_key))

            # Column A: Role label
            ws.cell(row=row, column=1, value=label).font = NORMAL_FONT
            ws.cell(row=row, column=1).border = THIN_BORDER

            # Column B: Allocation % (formula from Project Portfolio)
            alloc_cell_ref = f"$B${row}"
            cell = ws.cell(row=row, column=2)
            cell.value = f"=INDEX('Project Portfolio'!${port_col}:${port_col},MATCH($A${header_row},'Project Portfolio'!$A:$A,0))"
            cell.number_format = '0%'
            _style_data_cell(cell)

            # Columns C-H: role × phase hours = est × alloc% × role_phase_effort
            for i, phase in enumerate(SDLC_PHASES):
                col = 3 + i
                acol = PHASE_ASSUMPTION_COLS[phase]
                cell = ws.cell(row=row, column=col)
                # est_hours × allocation% × role_phase_effort
                cell.value = f"={est_cell}*B{row}*RM_Assumptions!{acol}{assump_row}"
                cell.number_format = '0.0'
                _style_data_cell(cell)

            # Column I: Total hours for this role = SUM
            cell = ws.cell(row=row, column=9)
            cell.value = f"=SUM(C{row}:H{row})"
            cell.number_format = '0.0'
            cell.font = BOLD_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

            # Column J: Days = total_role_hours / (per_person_capacity × 0.85) × 5
            cap_row = capacity_rows.get(role_key)
            if cap_row:
                cell = ws.cell(row=row, column=10)
                cell.value = f'=IFERROR(I{row}/($D${cap_row}*0.85)*5,0)'
                cell.number_format = '0.0'
                _style_data_cell(cell)

            row += 1

        # ── Totals row ──
        first_role_row = role_rows[0][0] if role_rows else row
        last_role_row = role_rows[-1][0] if role_rows else row

        ws.cell(row=row, column=1, value="TOTAL").font = BOLD_FONT
        ws.cell(row=row, column=1).fill = LIGHT_BLUE_FILL
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2, value="").fill = LIGHT_BLUE_FILL
        ws.cell(row=row, column=2).border = THIN_BORDER

        # Sum each phase column across roles
        for i in range(6):
            col = 3 + i
            cl = get_column_letter(col)
            cell = ws.cell(row=row, column=col)
            cell.value = f"=SUM({cl}{first_role_row}:{cl}{last_role_row})"
            cell.number_format = '0.0'
            cell.font = BOLD_FONT
            cell.fill = LIGHT_BLUE_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

        # Grand total
        cell = ws.cell(row=row, column=9)
        cell.value = f"=SUM(I{first_role_row}:I{last_role_row})"
        cell.number_format = '0.0'
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")
        total_row = row

        row += 1

        # ── Reconciliation row ──
        ws.cell(row=row, column=1, value="Reconciliation").font = BOLD_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER

        # Estimate
        cell = ws.cell(row=row, column=2)
        cell.value = f'="Est: "&TEXT({est_cell},"#,##0")&" hrs"'
        cell.border = THIN_BORDER

        # Allocated
        cell = ws.cell(row=row, column=3)
        cell.value = f'="Alloc: "&TEXT(I{total_row},"#,##0.0")&" hrs"'
        cell.border = THIN_BORDER

        # Gap
        cell = ws.cell(row=row, column=4)
        cell.value = f'=IF(ABS({est_cell}-I{total_row})<0.5,"RECONCILED","GAP: "&TEXT({est_cell}-I{total_row},"+#,##0.0;-#,##0.0")&" hrs")'
        cell.font = BOLD_FONT
        cell.border = THIN_BORDER
        # Conditional color (use formula result to decide — will show as text)

        row += 1

        # ── Phase Durations row (days per phase, bottleneck-driven) ──
        ws.cell(row=row, column=1, value="Phase Duration (days)").font = BOLD_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2, value="").border = THIN_BORDER

        # For each phase: MAX across roles of (role_phase_hours / (per_person_cap × 0.85)) × 5
        duration_cells = []
        for i, phase in enumerate(SDLC_PHASES):
            col = 3 + i
            cl = get_column_letter(col)

            # Build MAX formula: MAX(role1_phase_hrs/cap1, role2_phase_hrs/cap2, ...) × 5
            max_parts = []
            for rrow, rkey in role_rows:
                cap_row = capacity_rows.get(rkey)
                if cap_row:
                    max_parts.append(f"IFERROR({cl}{rrow}/($D${cap_row}*0.85),0)")

            if max_parts:
                max_formula = f'=MAX({",".join(max_parts)})*5'
            else:
                max_formula = "=0"

            cell = ws.cell(row=row, column=col)
            cell.value = max_formula
            cell.number_format = '0.0'
            cell.font = Font(bold=True, size=10)
            cell.fill = GOLD_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
            duration_cells.append(f"{cl}{row}")

        # Total duration (sum of phase durations)
        cell = ws.cell(row=row, column=9)
        cell.value = f'=SUM(C{row}:H{row})'
        cell.number_format = '0.0'
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

        # Label
        cell = ws.cell(row=row, column=10)
        cell.value = f'="Total: "&TEXT(I{row},"0.0")&" days"'
        cell.font = BOLD_FONT
        cell.border = THIN_BORDER

        row += 2  # gap before next project
        return row

    # ================================================================
    # CAPACITY SUMMARY — Supply vs Demand + Project List
    # ================================================================
    def _write_capacity_summary(self, wb):
        ws = wb.create_sheet("Capacity Summary")

        ws.merge_cells("A1:J1")
        ws["A1"] = "ETE IT PMO — Capacity Summary"
        ws["A1"].font = TITLE_FONT
        ws["A2"] = "All supply/headcount values are formulas — change Team Roster and this sheet recalculates"
        ws["A2"].font = SMALL_FONT

        roster = self._data["roster"]
        active = self._data["active_portfolio"]
        utilization = self.engine.compute_utilization()

        # Key numbers — formulas referencing the supply/demand table below
        row = 4
        kpi_labels = ["Team Size", "Active Projects", "Weekly Supply", "Weekly Demand", "Utilization"]
        for i, label in enumerate(kpi_labels):
            col_kpi = 1 + i * 2
            ws.cell(row=row, column=col_kpi, value=label).font = Font(bold=True, size=9, color="666666")

        # KPI values: Team Size = total headcount formula, Active Projects = computed count,
        # Supply/Demand/Util reference the totals row built below
        # We know the totals row will be at row 10 + len(active_roles) = row 10 + 8 = 18
        # But let's compute it precisely after building the table.
        # For now, store the KPI value row to fill in later.
        kpi_value_row = row + 1

        # Supply vs Demand by Role
        row = 8
        ws.cell(row=row, column=1, value="Supply vs. Demand by Role").font = SECTION_FONT
        row += 1
        _write_header_row(ws, row, [
            "Role", "Headcount", "Per-Person\n(hrs/wk)", "Team Supply\n(hrs/wk)",
            "Team Demand\n(hrs/wk)", "Gap\n(hrs/wk)", "Utilization", "Status"
        ])
        header_row_svd = row

        row += 1
        svd_first_row = row
        # Track rows for each role so we can build the total row
        svd_role_rows = {}

        for role_key, label, port_col, assump_row, roster_name in ROLE_CONFIG:
            u = utilization.get(role_key)
            if not u:
                continue

            svd_role_rows[role_key] = row

            ws.cell(row=row, column=1, value=label).font = BOLD_FONT
            ws.cell(row=row, column=1).border = THIN_BORDER

            # Col B: Headcount = COUNTIF on Team Roster
            cell = ws.cell(row=row, column=2)
            cell.value = f'=COUNTIF(\'Team Roster\'!$B$4:$B$26,"{roster_name}")'
            _style_data_cell(cell)

            # Col C: Per-Person capacity = AVERAGEIFS on Team Roster
            cell = ws.cell(row=row, column=3)
            cell.value = f'=IFERROR(AVERAGEIFS(\'Team Roster\'!$K$4:$K$26,\'Team Roster\'!$B$4:$B$26,"{roster_name}"),0)'
            cell.number_format = '0.0'
            _style_data_cell(cell)

            # Col D: Team Supply = SUMIFS on Team Roster
            cell = ws.cell(row=row, column=4)
            cell.value = f'=SUMIFS(\'Team Roster\'!$K$4:$K$26,\'Team Roster\'!$B$4:$B$26,"{roster_name}")'
            cell.number_format = '0.0'
            _style_data_cell(cell)

            # Col E: Demand — computed value (requires complex SDLC phase iteration across projects)
            demand_val = round(u.demand_hrs_week, 1)
            cell = ws.cell(row=row, column=5, value=demand_val)
            cell.number_format = '0.0'
            _style_data_cell(cell)

            # Col F: Gap = Supply - Demand (formula)
            cell = ws.cell(row=row, column=6)
            cell.value = f"=D{row}-E{row}"
            cell.number_format = '0.0'
            _style_data_cell(cell)

            # Col G: Utilization = Demand / Supply (formula)
            cell = ws.cell(row=row, column=7)
            cell.value = f'=IFERROR(E{row}/D{row},0)'
            cell.number_format = '0%'
            _style_data_cell(cell)

            # Col H: Status = formula based on utilization
            cell = ws.cell(row=row, column=8)
            cell.value = f'=IF(G{row}>=1,"RED",IF(G{row}>=0.8,"YELLOW","GREEN"))'
            _style_data_cell(cell)
            cell.font = BOLD_FONT

            # Conditional formatting via fill (static based on current data for visual appearance)
            ws.cell(row=row, column=6).fill = GREEN_FILL if u.supply_hrs_week - u.demand_hrs_week >= 0 else RED_FILL
            ws.cell(row=row, column=8).fill = _util_fill(u.utilization_pct)

            row += 1

        svd_last_row = row - 1

        # Total row (formulas)
        ws.cell(row=row, column=1, value="TOTAL").font = BOLD_FONT
        ws.cell(row=row, column=1).fill = LIGHT_BLUE_FILL
        ws.cell(row=row, column=1).border = THIN_BORDER

        for col_idx, col_letter in [(2, "B"), (4, "D"), (5, "E")]:
            cell = ws.cell(row=row, column=col_idx)
            cell.value = f"=SUM({col_letter}{svd_first_row}:{col_letter}{svd_last_row})"
            cell.font = BOLD_FONT
            cell.fill = LIGHT_BLUE_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
            if col_idx in (4, 5):
                cell.number_format = '0.0'

        # Total Gap = Total Supply - Total Demand
        cell = ws.cell(row=row, column=6)
        cell.value = f"=D{row}-E{row}"
        cell.font = BOLD_FONT
        cell.fill = LIGHT_BLUE_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")
        cell.number_format = '0.0'

        total_row = row

        # Now fill in KPI values with formulas referencing the totals row
        # Team Size
        cell = ws.cell(row=kpi_value_row, column=1, value=f"=B{total_row}")
        cell.font = Font(bold=True, size=18, color="4472C4")
        # Active Projects (count — static since it requires parsing portfolio status)
        cell = ws.cell(row=kpi_value_row, column=3, value=len(active))
        cell.font = Font(bold=True, size=18, color="4472C4")
        # Weekly Supply
        cell = ws.cell(row=kpi_value_row, column=5)
        cell.value = f'=TEXT(D{total_row},"0")&" hrs"'
        cell.font = Font(bold=True, size=18, color="4472C4")
        # Weekly Demand
        cell = ws.cell(row=kpi_value_row, column=7)
        cell.value = f'=TEXT(E{total_row},"0")&" hrs"'
        cell.font = Font(bold=True, size=18, color="4472C4")
        # Utilization
        cell = ws.cell(row=kpi_value_row, column=9)
        cell.value = f'=IFERROR(TEXT(E{total_row}/D{total_row},"0%"),"0%")'
        cell.font = Font(bold=True, size=18, color="4472C4")

        # Active Projects with duration estimates
        row = total_row + 3
        ws.cell(row=row, column=1, value="Active Projects — Bottom-Up Duration Estimates").font = SECTION_FONT
        row += 1
        _write_header_row(ws, row, [
            "ID", "Project", "Priority", "Est Hrs",
            "Allocated", "Duration\n(days)", "Bottleneck\nRole",
            "Start", "End"
        ])

        row += 1
        priority_order = {"Highest": 0, "High": 1, "Medium": 2, "Low": 3}
        for project in sorted(active, key=lambda p: (priority_order.get(p.priority, 9), p.name)):
            active_roles_p = {k: v for k, v in project.role_allocations.items() if v > 0}
            if not active_roles_p or project.est_hours <= 0:
                continue

            dur = self.engine.estimate_project_duration(project)
            bottlenecks = [p["bottleneck_role"] for p in dur["phases"] if p.get("bottleneck_role")]
            primary_bn = max(set(bottlenecks), key=bottlenecks.count) if bottlenecks else ""
            bn_label = ""
            for rk, rl, _, _, _ in ROLE_CONFIG:
                if rk == primary_bn:
                    bn_label = rl
                    break

            start_str = project.start_date.isoformat() if project.start_date else ""
            end_str = project.end_date.isoformat() if project.end_date else ""

            if not project.start_date:
                try:
                    dates = self.engine.suggest_dates(
                        project.est_hours, active_roles_p,
                        exclude_project_id=project.id
                    )
                    start_str = f"→ {dates.get('suggested_start', '')}"
                    end_str = f"→ {dates.get('suggested_end', '')}"
                except Exception:
                    start_str = "TBD"
                    end_str = "TBD"

            for col, val in [
                (1, project.id), (2, project.name), (3, project.priority),
                (4, project.est_hours), (5, dur["allocated_hours"]),
                (6, dur["total_duration_days"]),
                (7, bn_label), (8, start_str), (9, end_str),
            ]:
                cell = ws.cell(row=row, column=col, value=val)
                _style_data_cell(cell, center=(col != 2))
                if col == 1:
                    cell.font = BOLD_FONT

            if not dur["reconciled"]:
                ws.cell(row=row, column=5).fill = YELLOW_FILL
            if not project.start_date:
                ws.cell(row=row, column=8).font = Font(size=10, italic=True, color="4472C4")
                ws.cell(row=row, column=9).font = Font(size=10, italic=True, color="4472C4")

            row += 1

        for col, w in [(1, 24), (2, 14), (3, 14), (4, 14), (5, 14),
                        (6, 14), (7, 12), (8, 24), (9, 16)]:
            ws.column_dimensions[get_column_letter(col)].width = w

    # ================================================================
    # ROLE CAPACITY PLANNER — Scenario modeling + weekly role heatmap
    # ================================================================
    def _write_role_capacity_planner(self, wb):
        ws = wb.create_sheet("Role Capacity Planner")

        ws.merge_cells("A1:N1")
        ws["A1"] = "ETE IT PMO — Role Capacity Planner"
        ws["A1"].font = TITLE_FONT
        ws["A2"] = "Scenario table uses formulas — change Team Roster and scenarios recalculate"
        ws["A2"].font = SMALL_FONT

        active = self._data["active_portfolio"]
        roster = self._data["roster"]
        assumptions = self._data["assumptions"]
        phase_weights = assumptions.sdlc_phase_weights
        role_phase_efforts = assumptions.role_phase_efforts
        utilization = self.engine.compute_utilization()
        supply = self.engine.compute_supply_by_role()
        per_person = self.engine.compute_per_person_capacity()
        role_counts = Counter(m.role_key for m in roster)

        # ────────────────────────────────────────────────────────────
        # SECTION 1: Scenario Table — current + what-if headcount
        # ────────────────────────────────────────────────────────────
        row = 4
        ws.cell(row=row, column=1, value="Headcount Scenario Analysis").font = SECTION_FONT
        row += 1
        ws.cell(row=row, column=1,
                value="Shows current state and impact of adding +1, +2, or +3 people per role"
                ).font = SMALL_FONT
        row += 2

        _write_header_row(ws, row, [
            "Role", "Current\nHeadcount", "Per-Person\n(hrs/wk)",
            "Team Supply\n(hrs/wk)", "Demand\n(hrs/wk)", "Current\nUtil %",
            "Status",
            "+1 Person\nSupply", "+1 Person\nUtil %", "+1 Status",
            "+2 People\nSupply", "+2 People\nUtil %", "+2 Status",
            "+3 People\nSupply", "+3 People\nUtil %", "+3 Status",
        ])

        row += 1
        scenario_start_row = row

        for role_key, label, port_col, assump_row, roster_name in ROLE_CONFIG:
            u = utilization.get(role_key)
            if not u:
                continue

            ws.cell(row=row, column=1, value=label).font = BOLD_FONT
            ws.cell(row=row, column=1).border = THIN_BORDER

            # Col B: Headcount = COUNTIF on Team Roster (formula)
            cell = ws.cell(row=row, column=2)
            cell.value = f'=COUNTIF(\'Team Roster\'!$B$4:$B$26,"{roster_name}")'
            _style_data_cell(cell)

            # Col C: Per-Person capacity = AVERAGEIFS on Team Roster (formula)
            cell = ws.cell(row=row, column=3)
            cell.value = f'=IFERROR(AVERAGEIFS(\'Team Roster\'!$K$4:$K$26,\'Team Roster\'!$B$4:$B$26,"{roster_name}"),0)'
            cell.number_format = '0.0'
            _style_data_cell(cell)

            # Col D: Team Supply = SUMIFS on Team Roster (formula)
            cell = ws.cell(row=row, column=4)
            cell.value = f'=SUMIFS(\'Team Roster\'!$K$4:$K$26,\'Team Roster\'!$B$4:$B$26,"{roster_name}")'
            cell.number_format = '0.0'
            _style_data_cell(cell)

            # Col E: Demand — computed value (requires complex SDLC phase iteration)
            demand_val = round(u.demand_hrs_week, 1)
            cell = ws.cell(row=row, column=5, value=demand_val)
            cell.number_format = '0.0'
            _style_data_cell(cell)

            # Col F: Current Util % = Demand / Supply (formula)
            cell = ws.cell(row=row, column=6)
            cell.value = f'=IFERROR(E{row}/D{row},0)'
            cell.number_format = '0%'
            _style_data_cell(cell)
            cell.fill = _util_fill(u.utilization_pct)
            cell.font = BOLD_FONT

            # Col G: Status = formula based on utilization
            cell = ws.cell(row=row, column=7)
            cell.value = f'=IF(F{row}>=1,"OVER",IF(F{row}>=0.85,"AT LIMIT",IF(F{row}>=0.8,"TIGHT","OK")))'
            _style_data_cell(cell)
            cell.fill = _util_fill(u.utilization_pct)
            cell.font = BOLD_FONT

            # +1, +2, +3 scenarios (all formulas referencing current row's supply/demand/per-person)
            for offset, base_col in [(1, 8), (2, 11), (3, 14)]:
                # New Supply = Current Supply + Per-Person * N (formula)
                cell = ws.cell(row=row, column=base_col)
                cell.value = f"=D{row}+C{row}*{offset}"
                cell.number_format = '0.0'
                _style_data_cell(cell)

                # New Util % = Demand / New Supply (formula)
                new_supply_ref = f"{get_column_letter(base_col)}{row}"
                cell = ws.cell(row=row, column=base_col + 1)
                cell.value = f"=IFERROR(E{row}/{new_supply_ref},0)"
                cell.number_format = '0%'
                _style_data_cell(cell)
                # Static fill based on current data for visual appearance
                current_supply_val = u.supply_hrs_week
                pp_val = per_person.get(role_key, 0)
                new_supply_val = current_supply_val + (pp_val * offset)
                new_util_val = u.demand_hrs_week / new_supply_val if new_supply_val > 0 else 0
                cell.fill = _util_fill(new_util_val)
                cell.font = BOLD_FONT

                # New Status = formula
                util_ref = f"{get_column_letter(base_col + 1)}{row}"
                cell = ws.cell(row=row, column=base_col + 2)
                cell.value = f'=IF({util_ref}>=1,"OVER",IF({util_ref}>=0.85,"AT LIMIT",IF({util_ref}>=0.8,"TIGHT","OK")))'
                _style_data_cell(cell)
                cell.fill = _util_fill(new_util_val)
                cell.font = BOLD_FONT

            row += 1

        # ────────────────────────────────────────────────────────────
        # SECTION 2: Impact on Project Durations
        # ────────────────────────────────────────────────────────────
        row += 2
        ws.cell(row=row, column=1, value="Impact on Project Durations (days)").font = SECTION_FONT
        row += 1
        ws.cell(row=row, column=1,
                value="How project durations change if we add 1 person to a role (bottleneck relief)"
                ).font = SMALL_FONT
        row += 2

        # Header: Project | Current Days | +1 BA | +1 Func | +1 Tech | +1 Dev | +1 PM | ...
        duration_headers = ["ID", "Project", "Est Hrs", "Current\nDays"]
        scenario_roles = []
        for role_key, label, _, _, _ in ROLE_CONFIG:
            if role_counts.get(role_key, 0) > 0:
                short_label = label.split("/")[0].replace("IT ", "")
                duration_headers.append(f"+1 {short_label}\nDays")
                scenario_roles.append(role_key)

        _write_header_row(ws, row, duration_headers)
        row += 1

        priority_order = {"Highest": 0, "High": 1, "Medium": 2, "Low": 3}
        for project in sorted(active, key=lambda p: (priority_order.get(p.priority, 9), p.name)):
            active_roles = {k: v for k, v in project.role_allocations.items() if v > 0}
            if not active_roles or project.est_hours <= 0:
                continue

            # Current duration
            dur = self.engine.estimate_project_duration(project)
            current_days = dur["total_duration_days"]

            ws.cell(row=row, column=1, value=project.id).font = BOLD_FONT
            ws.cell(row=row, column=1).border = THIN_BORDER
            cell = ws.cell(row=row, column=2, value=project.name)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell = ws.cell(row=row, column=3, value=project.est_hours)
            _style_data_cell(cell)
            cell = ws.cell(row=row, column=4, value=current_days)
            cell.font = BOLD_FONT
            _style_data_cell(cell)

            # Scenario durations: simulate adding 1 person to each role
            col = 5
            for scenario_role in scenario_roles:
                # Recalculate per-person capacity as if we added 1 person to this role
                original_pp = per_person.get(scenario_role, 0)
                original_hc = role_counts.get(scenario_role, 0)

                if original_hc > 0 and original_pp > 0:
                    # New per-person stays same (it's average), but supply increases
                    # Duration impact: if this role is the bottleneck, adding a person
                    # means the project could have 2 people on that role
                    # Simplified: new capacity for that role = original + per_person
                    # This only helps if role is actually on this project

                    # Recalculate duration with increased capacity
                    new_per_person = dict(per_person)
                    # Adding 1 person doesn't change per-person avg much,
                    # but it means concurrent capacity increases.
                    # For bottleneck calc: the role gets more throughput
                    # Simulate by reducing hours needed (as if 2 people share work)
                    if scenario_role in active_roles:
                        # This role is on the project — adding a person helps
                        # New effective per-person for this role on this project:
                        # 2 people can split the work, so effective capacity doubles
                        boosted_pp = dict(per_person)
                        boosted_pp[scenario_role] = original_pp * 2  # 2 people now

                        # Recalculate phase durations with boosted capacity
                        total_days_new = 0
                        for phase in SDLC_PHASES:
                            max_weeks = 0
                            for rk, alloc in active_roles.items():
                                if rk not in role_phase_efforts:
                                    continue
                                effort = role_phase_efforts[rk].get(phase, 0)
                                hrs = project.est_hours * alloc * effort
                                cap = boosted_pp.get(rk, 0) * 0.85
                                if cap > 0 and hrs > 0:
                                    weeks = hrs / cap
                                    if weeks > max_weeks:
                                        max_weeks = weeks
                            total_days_new += max_weeks * 5

                        new_days = round(total_days_new, 1)
                    else:
                        new_days = current_days  # role not on project, no change

                    cell = ws.cell(row=row, column=col, value=new_days)
                    _style_data_cell(cell)

                    # Highlight improvement
                    if new_days < current_days - 0.5:
                        cell.fill = GREEN_FILL
                        cell.font = Font(bold=True, size=10, color="006100")
                    elif new_days < current_days:
                        cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                else:
                    ws.cell(row=row, column=col, value=current_days).border = THIN_BORDER

                col += 1

            row += 1

        # ────────────────────────────────────────────────────────────
        # SECTION 3: Weekly Role Heatmap
        # ────────────────────────────────────────────────────────────
        row += 2
        ws.cell(row=row, column=1, value="Weekly Role Utilization Heatmap").font = SECTION_FONT
        row += 1
        ws.cell(row=row, column=1,
                value="Role-level utilization by week — shows when role groups get hot"
                ).font = SMALL_FONT
        row += 1

        # Legend
        for col, (label, fill) in enumerate([
            ("Overloaded (>100%)", RED_FILL),
            ("At Capacity (80-100%)", YELLOW_FILL),
            ("Healthy (<80%)", GREEN_FILL),
        ], 1):
            ws.cell(row=row, column=col, value=label).font = Font(bold=True, size=9)
            ws.cell(row=row, column=col).fill = fill
        row += 1

        today = date.today()
        days_to_monday = (7 - today.weekday()) % 7
        start_date = today + timedelta(days=days_to_monday if days_to_monday else 0)
        num_weeks = 26

        # Build weekly demand by role
        role_weekly_demand = defaultdict(lambda: defaultdict(float))

        for project in active:
            if not project.start_date or not project.end_date or project.est_hours <= 0:
                continue
            proj_days = (project.end_date - project.start_date).days
            if proj_days <= 0:
                continue

            phase_bounds = []
            cum = 0
            for phase in SDLC_PHASES:
                w = phase_weights.get(phase, 0.0)
                pd = round(proj_days * w)
                phase_bounds.append((phase, cum, cum + pd))
                cum += pd

            for role_key, alloc_pct in project.role_allocations.items():
                if alloc_pct <= 0 or role_key not in role_phase_efforts:
                    continue

                role_hrs = project.est_hours * alloc_pct
                proj_weeks = max(1, proj_days / 7.0)

                for week_idx in range(num_weeks):
                    ws_date = start_date + timedelta(weeks=week_idx)
                    we_date = ws_date + timedelta(days=7)
                    if we_date <= project.start_date or ws_date >= project.end_date:
                        continue

                    day_off = max(0, (ws_date - project.start_date).days)
                    cur_phase = SDLC_PHASES[-1]
                    for pn, ps, pe in phase_bounds:
                        if ps <= day_off < pe:
                            cur_phase = pn
                            break

                    effort = role_phase_efforts[role_key].get(cur_phase, 0.0)
                    weekly_hrs = role_hrs * effort / proj_weeks
                    role_weekly_demand[role_key][week_idx] += weekly_hrs

        # Headers
        row += 1
        for col, label in [(1, "Role"), (2, "Headcount"), (3, "Supply\n(hrs/wk)")]:
            cell = ws.cell(row=row, column=col, value=label)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = THIN_BORDER

        for week_idx in range(num_weeks):
            week_date = start_date + timedelta(weeks=week_idx)
            col = 4 + week_idx
            cell = ws.cell(row=row, column=col, value=week_date.strftime("%m/%d"))
            cell.font = Font(bold=True, size=8, color="FFFFFF")
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER
            ws.column_dimensions[get_column_letter(col)].width = 6.5

        row += 1

        for role_key, label, _, _, _ in ROLE_CONFIG:
            role_supply = supply.get(role_key, 0)
            if role_supply <= 0 and role_key not in role_weekly_demand:
                continue

            ws.cell(row=row, column=1, value=label).font = BOLD_FONT
            ws.cell(row=row, column=1).border = THIN_BORDER
            ws.cell(row=row, column=2, value=role_counts.get(role_key, 0))
            ws.cell(row=row, column=2).border = THIN_BORDER
            ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=3, value=round(role_supply, 1))
            ws.cell(row=row, column=3).border = THIN_BORDER
            ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")

            for week_idx in range(num_weeks):
                col = 4 + week_idx
                demand = role_weekly_demand[role_key][week_idx]
                util = demand / role_supply if role_supply > 0 else 0

                cell = ws.cell(row=row, column=col)
                cell.value = round(util * 100)
                cell.number_format = '0"%"'
                cell.fill = _util_fill(util)
                cell.font = Font(size=8, bold=(util >= 0.80))
                cell.alignment = Alignment(horizontal="center")
                cell.border = THIN_BORDER

            row += 1

        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 12

    # ================================================================
    # CAPACITY HEATMAP — Per-person weekly utilization
    # ================================================================
    def _write_capacity_heatmap(self, wb):
        ws = wb.create_sheet("Capacity Heatmap")

        ws.merge_cells("A1:AD1")
        ws["A1"] = "ETE IT PMO — Weekly Capacity Heatmap (Per Person)"
        ws["A1"].font = TITLE_FONT

        for col, (label, fill) in enumerate([
            ("Overloaded (>100%)", RED_FILL),
            ("At Capacity (80-100%)", YELLOW_FILL),
            ("Healthy (<80%)", GREEN_FILL),
        ], 1):
            ws.cell(row=2, column=col, value=label).font = Font(bold=True, size=9)
            ws.cell(row=2, column=col).fill = fill

        active = self._data["active_portfolio"]
        roster = self._data["roster"]
        assumptions = self._data["assumptions"]
        phase_weights = assumptions.sdlc_phase_weights
        role_phase_efforts = assumptions.role_phase_efforts

        today = date.today()
        days_to_monday = (7 - today.weekday()) % 7
        start_date = today + timedelta(days=days_to_monday if days_to_monday else 0)
        num_weeks = 26

        person_demand = defaultdict(lambda: defaultdict(float))

        for project in active:
            if not project.start_date or not project.end_date or project.est_hours <= 0:
                continue
            proj_days = (project.end_date - project.start_date).days
            if proj_days <= 0:
                continue

            phase_bounds = []
            cum = 0
            for phase in SDLC_PHASES:
                w = phase_weights.get(phase, 0.0)
                pd = round(proj_days * w)
                phase_bounds.append((phase, cum, cum + pd))
                cum += pd

            for role_key, alloc_pct in project.role_allocations.items():
                if alloc_pct <= 0 or role_key not in role_phase_efforts:
                    continue
                role_members = [m for m in roster if m.role_key == role_key]
                if not role_members:
                    continue

                role_hrs = project.est_hours * alloc_pct
                proj_weeks = max(1, proj_days / 7.0)
                per_member_share = 1.0 / len(role_members)

                for week_idx in range(num_weeks):
                    ws_date = start_date + timedelta(weeks=week_idx)
                    we_date = ws_date + timedelta(days=7)
                    if we_date <= project.start_date or ws_date >= project.end_date:
                        continue

                    day_off = max(0, (ws_date - project.start_date).days)
                    cur_phase = SDLC_PHASES[-1]
                    for pn, ps, pe in phase_bounds:
                        if ps <= day_off < pe:
                            cur_phase = pn
                            break

                    effort = role_phase_efforts[role_key].get(cur_phase, 0.0)
                    weekly_hrs = role_hrs * effort / proj_weeks

                    for member in role_members:
                        person_demand[member.name][week_idx] += weekly_hrs * per_member_share

        row = 4
        for col, label in [(1, "Resource"), (2, "Role"), (3, "Team"), (4, "Capacity\n(hrs/wk)")]:
            cell = ws.cell(row=row, column=col, value=label)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = THIN_BORDER

        for week_idx in range(num_weeks):
            week_date = start_date + timedelta(weeks=week_idx)
            col = 5 + week_idx
            cell = ws.cell(row=row, column=col, value=week_date.strftime("%m/%d"))
            cell.font = Font(bold=True, size=8, color="FFFFFF")
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER
            ws.column_dimensions[get_column_letter(col)].width = 6.5

        row = 5
        role_order_keys = [r[0] for r in ROLE_CONFIG]
        for member in sorted(roster, key=lambda m: (role_order_keys.index(m.role_key)
                              if m.role_key in role_order_keys else 99, m.name)):
            ws.cell(row=row, column=1, value=member.name).font = NORMAL_FONT
            ws.cell(row=row, column=1).border = THIN_BORDER
            ws.cell(row=row, column=2, value=member.role).font = NORMAL_FONT
            ws.cell(row=row, column=2).border = THIN_BORDER
            ws.cell(row=row, column=3, value=member.team or "").font = NORMAL_FONT
            ws.cell(row=row, column=3).border = THIN_BORDER
            ws.cell(row=row, column=4, value=round(member.project_capacity_hrs, 1))
            ws.cell(row=row, column=4).border = THIN_BORDER
            ws.cell(row=row, column=4).alignment = Alignment(horizontal="center")

            capacity = member.project_capacity_hrs
            for week_idx in range(num_weeks):
                col = 5 + week_idx
                demand = person_demand[member.name][week_idx]
                util = demand / capacity if capacity > 0 else 0

                cell = ws.cell(row=row, column=col)
                cell.value = round(util * 100)
                cell.number_format = '0"%"'
                cell.fill = _util_fill(util)
                cell.font = Font(size=8, bold=(util >= 0.80))
                cell.alignment = Alignment(horizontal="center")
                cell.border = THIN_BORDER

            row += 1

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 10

    # ================================================================
    # RM_GANTT — Visual Gantt chart
    # ================================================================
    def _write_gantt(self, wb):
        ws = wb.create_sheet("RM_Gantt")

        ws.merge_cells("A1:AZ1")
        ws["A1"] = "ETE IT PMO — Project Gantt Chart"
        ws["A1"].font = TITLE_FONT
        ws["A2"] = f"Generated: {date.today().isoformat()}  |  Color = priority"
        ws["A2"].font = SMALL_FONT

        active = self._data["active_portfolio"]
        priority_order = {"Highest": 0, "High": 1, "Medium": 2, "Low": 3}
        active = sorted(active, key=lambda p: (priority_order.get(p.priority, 9), p.name))

        all_starts = [p.start_date for p in active if p.start_date]
        all_ends = [p.end_date for p in active if p.end_date]

        if not all_starts:
            today = date.today()
            gantt_start = today
            gantt_end = today + timedelta(weeks=26)
        else:
            gantt_start = min(all_starts)
            gantt_end = max(all_ends) + timedelta(weeks=4) if all_ends else gantt_start + timedelta(weeks=26)

        gantt_start = gantt_start - timedelta(days=gantt_start.weekday())
        num_weeks = min(max(1, (gantt_end - gantt_start).days // 7 + 1), 52)

        row = 4
        _write_header_row(ws, row, ["ID", "Project Name", "Priority", "Est Hrs", "Start", "End", "Dur"])

        for w in range(num_weeks):
            week_date = gantt_start + timedelta(weeks=w)
            col = 8 + w
            cell = ws.cell(row=row, column=col, value=week_date.strftime("%m/%d"))
            cell.font = Font(bold=True, size=7, color="FFFFFF")
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER
            ws.column_dimensions[get_column_letter(col)].width = 4

        today = date.today()
        current_week_idx = max(0, (today - gantt_start).days // 7)
        if current_week_idx < num_weeks:
            cell = ws.cell(row=3, column=8 + current_week_idx, value="NOW")
            cell.font = Font(bold=True, size=8, color="FF0000")
            cell.alignment = Alignment(horizontal="center")

        row = 5
        for project in active:
            if project.est_hours <= 0:
                continue

            ws.cell(row=row, column=1, value=project.id).font = BOLD_FONT
            ws.cell(row=row, column=1).border = THIN_BORDER
            ws.cell(row=row, column=2, value=project.name).font = NORMAL_FONT
            ws.cell(row=row, column=2).border = THIN_BORDER
            ws.cell(row=row, column=3, value=project.priority).border = THIN_BORDER
            ws.cell(row=row, column=4, value=project.est_hours).border = THIN_BORDER

            start = project.start_date
            end = project.end_date
            dur_str = ""

            if start and end:
                dur_str = f"{max(1, (end - start).days / 7):.0f}w"
            else:
                active_roles_p = {k: v for k, v in project.role_allocations.items() if v > 0}
                if active_roles_p:
                    try:
                        dates = self.engine.suggest_dates(
                            project.est_hours, active_roles_p,
                            exclude_project_id=project.id
                        )
                        if dates.get("suggested_start"):
                            start = date.fromisoformat(dates["suggested_start"])
                            end = date.fromisoformat(dates["suggested_end"])
                            dur_str = f"{dates['duration_days']:.0f}d (est)"
                    except Exception:
                        pass

            ws.cell(row=row, column=5, value=start.isoformat() if start else "TBD").border = THIN_BORDER
            ws.cell(row=row, column=6, value=end.isoformat() if end else "TBD").border = THIN_BORDER
            ws.cell(row=row, column=7, value=dur_str).border = THIN_BORDER

            for c in range(1, 8):
                ws.cell(row=row, column=c).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=2).alignment = Alignment(horizontal="left")

            if start and end:
                bar_fill = PRIORITY_FILLS.get(project.priority, LIGHT_BLUE_FILL)
                for w in range(num_weeks):
                    week_start = gantt_start + timedelta(weeks=w)
                    week_end = week_start + timedelta(days=6)
                    col = 8 + w
                    if week_end >= start and week_start <= end:
                        cell = ws.cell(row=row, column=col)
                        cell.fill = bar_fill
                        cell.border = THIN_BORDER

            row += 1

        row += 1
        ws.cell(row=row, column=1, value="Legend:").font = BOLD_FONT
        for i, (priority, fill) in enumerate(PRIORITY_FILLS.items()):
            ws.cell(row=row, column=2 + i * 2, value=priority).font = Font(bold=True, size=9)
            ws.cell(row=row, column=2 + i * 2).fill = fill

        for col, w in [(1, 10), (2, 38), (3, 10), (4, 8), (5, 12), (6, 12), (7, 10)]:
            ws.column_dimensions[get_column_letter(col)].width = w


if __name__ == "__main__":
    gen = DashboardGenerator()
    path = gen.generate_all()
    gen.connector.close()
    print(f"\nDone! Open: {path}")
