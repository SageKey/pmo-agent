"""
Import Synnergie timesheet data from Excel into the PMO database.
Reads from the Synnergie_Timesheets_0326.xlsx file and populates:
  - vendor_consultants
  - vendor_timesheets
  - approved_work
  - vendor_invoices (March 2026 invoice)

Usage: python import_synnergie.py [path_to_xlsx]
"""

import sys
from datetime import datetime
from pathlib import Path

# Ensure project root on path
sys.path.insert(0, str(Path(__file__).parent))

from sqlite_connector import SQLiteConnector
from data_layer import DB_PATH


# Excel sheet name → (canonical full name, billing info)
# Full names match team_members.name exactly
CONSULTANT_BILLING = {
    "Ajay":     {"full_name": "Ajay Kumar",       "billing_type": "MSA", "hourly_rate": 0.0,   "role_key": "functional"},
    "Bhavya":   {"full_name": "Bhavya Reddy",     "billing_type": "T&M", "hourly_rate": 60.0,  "role_key": "technical"},
    "Deepak":   {"full_name": "Deepak Gudwani",   "billing_type": "MSA", "hourly_rate": 0.0,   "role_key": "functional"},
    "Ravi":     {"full_name": "Ravindra Reddy",    "billing_type": "MSA", "hourly_rate": 0.0,   "role_key": "technical"},
    "Sangam":   {"full_name": "Sangamesh Koti",    "billing_type": "MSA", "hourly_rate": 0.0,   "role_key": "technical"},
    "Sarath":   {"full_name": "Sarath Yeturu",     "billing_type": "T&M", "hourly_rate": 200.0, "role_key": "technical"},
    "Vinod":    {"full_name": "Vinod Bollepally",   "billing_type": "MSA", "hourly_rate": 0.0,   "role_key": "dba"},
    "Vishnu":   {"full_name": "Vishnu Premen",     "billing_type": "MSA", "hourly_rate": 0.0,   "role_key": "technical"},
    "Akhilesh": {"full_name": "Akhilesh Mishra",   "billing_type": "T&M", "hourly_rate": 65.0,  "role_key": "technical"},
}


def import_data(xlsx_path: str):
    """Import all Synnergie data from the Excel file."""
    try:
        import openpyxl
    except ImportError:
        print("Error: openpyxl required. Install with: pip install openpyxl")
        return

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    connector = SQLiteConnector(DB_PATH)

    try:
        conn = connector._open()

        # ---------------------------------------------------------------
        # 1. Vendor Consultants
        # ---------------------------------------------------------------
        print("Importing vendor consultants...")
        for sheet_name, info in CONSULTANT_BILLING.items():
            connector.save_vendor_consultant({
                "name": info["full_name"],
                "billing_type": info["billing_type"],
                "hourly_rate": info["hourly_rate"],
                "role_key": info["role_key"],
                "active": 1,
            })
        print(f"  {len(CONSULTANT_BILLING)} consultants upserted.")

        # Build consultant ID lookup: full_name → id
        consultants = {c["name"]: c["id"] for c in connector.read_vendor_consultants()}
        # Also build sheet_name → consultant_id mapping
        sheet_to_id = {}
        for sheet_name, info in CONSULTANT_BILLING.items():
            cid = consultants.get(info["full_name"])
            if cid:
                sheet_to_id[sheet_name] = cid

        # ---------------------------------------------------------------
        # 2. Timesheet Entries
        # ---------------------------------------------------------------
        print("Importing timesheet entries...")
        sheet_names = list(CONSULTANT_BILLING.keys())
        total_entries = 0

        for sheet_name in sheet_names:
            if sheet_name not in wb.sheetnames:
                print(f"  WARNING: Sheet '{sheet_name}' not found, skipping.")
                continue

            ws = wb[sheet_name]
            consultant_id = sheet_to_id.get(sheet_name)
            if not consultant_id:
                print(f"  WARNING: Consultant '{sheet_name}' not mapped, skipping.")
                continue

            entries = 0
            for row in ws.iter_rows(min_row=7, max_row=100, values_only=True):
                entry_date = row[0]
                if not entry_date or not isinstance(entry_date, datetime):
                    continue

                hours = row[6]
                if not hours or hours == 0:
                    continue

                project_key = row[2]  # may be None
                project_name = row[3]  # may be None
                task_desc = row[4]
                work_type = row[5] or "Support"
                notes = row[7]

                # Normalize work_type
                if isinstance(work_type, str):
                    wt_lower = work_type.strip().lower()
                    work_type = "Project" if wt_lower == "project" else "Support"
                else:
                    work_type = "Support"

                connector.save_timesheet_entry({
                    "consultant_id": consultant_id,
                    "entry_date": entry_date.strftime("%Y-%m-%d"),
                    "project_key": project_key,
                    "project_name": project_name,
                    "task_description": task_desc,
                    "work_type": work_type,
                    "hours": float(hours),
                    "notes": notes,
                })
                entries += 1

            total_entries += entries
            print(f"  {sheet_name}: {entries} entries")

        print(f"  Total: {total_entries} timesheet entries imported.")

        # ---------------------------------------------------------------
        # 3. Approved Work Register
        # ---------------------------------------------------------------
        print("Importing approved work register...")
        if "Approved Work" in wb.sheetnames:
            ws = wb["Approved Work"]
            count = 0
            for row in ws.iter_rows(min_row=3, max_row=60, values_only=True):
                jira_key = row[0]
                if not jira_key:
                    continue

                title = row[1] or ""
                work_type = row[2]
                classification = row[3]
                approved_date = row[4]
                approver = row[5]
                notes = row[6]

                date_str = None
                if approved_date and isinstance(approved_date, datetime):
                    date_str = approved_date.strftime("%Y-%m-%d")

                connector.save_approved_work({
                    "jira_key": str(jira_key),
                    "title": str(title),
                    "work_type": str(work_type) if work_type else None,
                    "work_classification": str(classification) if classification else None,
                    "approved_date": date_str,
                    "approver": str(approver) if approver else None,
                    "notes": str(notes) if notes else None,
                })
                count += 1

            print(f"  {count} approved work items imported.")

        # ---------------------------------------------------------------
        # 4. March 2026 Invoice
        # ---------------------------------------------------------------
        print("Importing March 2026 invoice...")
        connector.save_invoice({
            "month": "2026-03",
            "msa_amount": 50000.0,
            "tm_amount": 51680.0,
            "total_amount": 101680.0,
            "invoice_number": "INV-ETEMar26",
            "received_date": "2026-04-02",
            "paid": 0,
            "notes": "MSA flat fee + T&M: Bhavya 176h@$60, Sarath 105h@$200, Akhilesh 72h@$65",
        })
        print("  Invoice recorded.")

        print("\nImport complete!")

    finally:
        connector.close()


if __name__ == "__main__":
    xlsx = sys.argv[1] if len(sys.argv) > 1 else "/Users/brettanderson/Downloads/Synnergie_Timesheets_0326.xlsx"
    if not Path(xlsx).exists():
        print(f"File not found: {xlsx}")
        sys.exit(1)
    import_data(xlsx)
