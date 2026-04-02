"""
Excel Connector for ETE IT PMO Resource Planning Agent.
Reads Project Portfolio, Team Roster, and RM_Assumptions sheets
from the ETE PMO workbook using openpyxl.
"""

from dataclasses import dataclass, field
from datetime import datetime, date
from pathlib import Path
from typing import Optional

import openpyxl

DEFAULT_WORKBOOK = "ETE PMO Resource Planner.xlsx"

# --- Role name mapping ---
# Maps Portfolio column names → canonical role keys used across the system.
# These also map to RM_Assumptions role labels and Roster role names.

PORTFOLIO_ROLE_COLUMNS = {
    # col_letter: (col_index, canonical_role_key)
    "U": (21, "ba"),
    "V": (22, "functional"),
    "W": (23, "technical"),
    "X": (24, "developer"),
    "Y": (25, "infrastructure"),
    "Z": (26, "dba"),
    "AA": (27, "pm"),
    "AB": (28, "wms"),
}

# RM_Assumptions role labels → canonical keys
ASSUMPTIONS_ROLE_MAP = {
    "PM": "pm",
    "DBA": "dba",
    "IT Business Analyst/SME": "ba",
    "IT Functional Analyst": "functional",
    "IT Tech Analyst/Developer": "technical",
    "Developer": "developer",
    "Infrastructure": "infrastructure",
    "WMS Consultant": "wms",
}

# Roster role names → canonical keys
ROSTER_ROLE_MAP = {
    "Project Manager": "pm",
    "DBA": "dba",
    "Business Analyst": "ba",
    "Functional": "functional",
    "Technical": "technical",
    "Developer": "developer",
    "Infrastructure": "infrastructure",
    "WMS Consultant": "wms",
}

SDLC_PHASES = ["discovery", "planning", "design", "build", "test", "deploy"]


@dataclass
class ProjectAssignment:
    """One person assigned to one project with a time allocation."""
    project_id: str
    person_name: str
    role_key: str
    allocation_pct: float  # 0.0-1.0, fraction of person's project capacity


@dataclass
class Project:
    id: str
    name: str
    type: Optional[str]
    portfolio: Optional[str]
    sponsor: Optional[str]
    health: Optional[str]
    pct_complete: float
    priority: Optional[str]
    start_date: Optional[date]
    end_date: Optional[date]
    actual_end: Optional[date]
    team: Optional[str]
    pm: Optional[str]
    ba: Optional[str]
    functional_lead: Optional[str]
    technical_lead: Optional[str]
    developer_lead: Optional[str]
    tshirt_size: Optional[str]
    est_hours: float
    est_cost: Optional[float]
    role_allocations: dict  # canonical_role_key → float (0.0-1.0)
    notes: Optional[str]
    sort_order: Optional[int]

    @property
    def is_active(self) -> bool:
        """Exclude POSTPONED and 100% complete projects."""
        if self.health and "POSTPONED" in self.health:
            return False
        if self.pct_complete >= 1.0:
            return False
        return True

    @property
    def duration_weeks(self) -> Optional[float]:
        """Project duration in weeks. None if dates missing."""
        if not self.start_date or not self.end_date:
            return None
        delta = self.end_date - self.start_date
        weeks = delta.days / 7.0
        return max(weeks, 1.0)  # floor at 1 week


@dataclass
class TeamMember:
    name: str
    role: str  # raw role name from roster
    role_key: str  # canonical role key
    team: Optional[str]
    vendor: Optional[str]
    classification: Optional[str]
    rate_per_hour: float
    weekly_hrs_available: float
    support_reserve_pct: float
    project_capacity_pct: float
    project_capacity_hrs: float


@dataclass
class RMAssumptions:
    base_hours_per_week: float  # 40
    admin_pct: float  # 0.10
    breakfix_pct: float  # 0.10
    project_pct: float  # 0.80
    available_project_hrs: float  # 32
    max_projects_per_person: int  # 3
    sdlc_phase_weights: dict  # phase_name → weight (sums to 1.0)
    role_phase_efforts: dict  # canonical_role_key → {phase_name → effort %}
    role_avg_efforts: dict  # canonical_role_key → avg effort across phases
    supply_by_role: dict  # canonical_role_key → {headcount, gross_hrs, project_hrs}


def _to_date(val) -> Optional[date]:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    return None


def _to_float(val, default=0.0) -> float:
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def _to_int(val, default=0) -> int:
    if val is None:
        return default
    try:
        return int(val)
    except (ValueError, TypeError):
        return default


class ExcelConnector:
    """Reads the ETE PMO workbook and returns structured data."""

    def __init__(self, workbook_path: Optional[str] = None):
        if workbook_path is None:
            workbook_path = str(Path(__file__).parent / DEFAULT_WORKBOOK)
        self.workbook_path = workbook_path
        self._wb = None

    def _open(self):
        if self._wb is None:
            self._wb = openpyxl.load_workbook(self.workbook_path, data_only=True)
        return self._wb

    def close(self):
        if self._wb:
            self._wb.close()
            self._wb = None

    @property
    def file_modified_time(self) -> datetime:
        return datetime.fromtimestamp(Path(self.workbook_path).stat().st_mtime)

    # ------------------------------------------------------------------
    # Project Portfolio
    # ------------------------------------------------------------------
    def read_portfolio(self) -> list[Project]:
        """Read Project Portfolio sheet (A3:AC42). Row 3 = headers, rows 4-42 = data."""
        wb = self._open()
        ws = wb["Project Portfolio"]
        projects = []

        for row in range(4, 43):
            pid = ws.cell(row=row, column=1).value
            name = ws.cell(row=row, column=2).value
            if not pid or not name:
                continue

            # Build role allocations dict
            role_allocs = {}
            for _col_letter, (col_idx, role_key) in PORTFOLIO_ROLE_COLUMNS.items():
                val = _to_float(ws.cell(row=row, column=col_idx).value, 0.0)
                role_allocs[role_key] = val

            project = Project(
                id=str(pid),
                name=str(name),
                type=ws.cell(row=row, column=3).value,
                portfolio=ws.cell(row=row, column=4).value,
                sponsor=ws.cell(row=row, column=5).value,
                health=ws.cell(row=row, column=6).value,
                pct_complete=_to_float(ws.cell(row=row, column=7).value, 0.0),
                priority=ws.cell(row=row, column=8).value,
                start_date=_to_date(ws.cell(row=row, column=9).value),
                end_date=_to_date(ws.cell(row=row, column=10).value),
                actual_end=_to_date(ws.cell(row=row, column=11).value),
                team=ws.cell(row=row, column=12).value,
                pm=ws.cell(row=row, column=13).value,
                ba=ws.cell(row=row, column=14).value,
                functional_lead=ws.cell(row=row, column=15).value,
                technical_lead=ws.cell(row=row, column=16).value,
                developer_lead=ws.cell(row=row, column=17).value,
                tshirt_size=ws.cell(row=row, column=18).value,
                est_hours=_to_float(ws.cell(row=row, column=19).value, 0.0),
                est_cost=_to_float(ws.cell(row=row, column=20).value, None),
                role_allocations=role_allocs,
                notes=ws.cell(row=row, column=29).value,
                sort_order=_to_int(ws.cell(row=row, column=30).value, None),
            )
            projects.append(project)

        return projects

    def read_active_portfolio(self) -> list[Project]:
        """Return only active projects (not POSTPONED, not 100% complete)."""
        return [p for p in self.read_portfolio() if p.is_active]

    # ------------------------------------------------------------------
    # Team Roster
    # ------------------------------------------------------------------
    def read_roster(self) -> list[TeamMember]:
        """Read Team Roster sheet (A3:K26). Row 3 = headers, rows 4-26 = data."""
        wb = self._open()
        ws = wb["Team Roster"]
        members = []

        for row in range(4, 27):
            name = ws.cell(row=row, column=1).value
            if not name or name == "TOTALS":
                continue

            raw_role = ws.cell(row=row, column=2).value or ""
            role_key = ROSTER_ROLE_MAP.get(raw_role, raw_role.lower())

            member = TeamMember(
                name=str(name),
                role=raw_role,
                role_key=role_key,
                team=ws.cell(row=row, column=3).value,
                vendor=ws.cell(row=row, column=4).value,
                classification=ws.cell(row=row, column=5).value,
                rate_per_hour=_to_float(ws.cell(row=row, column=6).value, 0.0),
                weekly_hrs_available=_to_float(ws.cell(row=row, column=7).value, 0.0),
                support_reserve_pct=_to_float(ws.cell(row=row, column=8).value, 0.0),
                project_capacity_pct=_to_float(ws.cell(row=row, column=9).value, 0.0),
                project_capacity_hrs=_to_float(ws.cell(row=row, column=11).value, 0.0),
            )
            members.append(member)

        return members

    # ------------------------------------------------------------------
    # RM_Assumptions
    # ------------------------------------------------------------------
    def read_assumptions(self) -> RMAssumptions:
        """Read RM_Assumptions sheet — time allocation, SDLC weights, role efforts, supply."""
        wb = self._open()
        ws = wb["RM_Assumptions"]

        # Time allocation (rows 6-12)
        base_hrs = _to_float(ws.cell(row=6, column=2).value, 40.0)
        admin_pct = _to_float(ws.cell(row=7, column=2).value, 0.10)
        breakfix_pct = _to_float(ws.cell(row=8, column=2).value, 0.10)
        project_pct = _to_float(ws.cell(row=9, column=2).value, 0.80)
        avail_hrs = _to_float(ws.cell(row=10, column=2).value, 32.0)
        max_projects = _to_int(ws.cell(row=12, column=2).value, 3)

        # SDLC phase weights (row 16, cols B-G)
        phase_weights = {}
        for i, phase in enumerate(SDLC_PHASES):
            phase_weights[phase] = _to_float(ws.cell(row=16, column=2 + i).value, 0.0)

        # Role effort by SDLC phase (rows 21-28, cols A-H)
        role_phase_efforts = {}
        role_avg_efforts = {}
        for row in range(21, 29):
            role_label = ws.cell(row=row, column=1).value
            if not role_label:
                continue
            role_key = ASSUMPTIONS_ROLE_MAP.get(role_label)
            if not role_key:
                continue

            efforts = {}
            for i, phase in enumerate(SDLC_PHASES):
                efforts[phase] = _to_float(ws.cell(row=row, column=2 + i).value, 0.0)
            role_phase_efforts[role_key] = efforts
            role_avg_efforts[role_key] = _to_float(ws.cell(row=row, column=8).value, 0.0)

        # Supply by role (rows 33-40)
        supply_by_role = {}
        for row in range(33, 41):
            role_label = ws.cell(row=row, column=1).value
            if not role_label or role_label == "TOTAL":
                continue
            role_key = ASSUMPTIONS_ROLE_MAP.get(role_label)
            if not role_key:
                continue
            supply_by_role[role_key] = {
                "headcount": _to_int(ws.cell(row=row, column=2).value, 0),
                "gross_hrs_week": _to_float(ws.cell(row=row, column=3).value, 0.0),
                "project_hrs_week": _to_float(ws.cell(row=row, column=4).value, 0.0),
            }

        return RMAssumptions(
            base_hours_per_week=base_hrs,
            admin_pct=admin_pct,
            breakfix_pct=breakfix_pct,
            project_pct=project_pct,
            available_project_hrs=avail_hrs,
            max_projects_per_person=max_projects,
            sdlc_phase_weights=phase_weights,
            role_phase_efforts=role_phase_efforts,
            role_avg_efforts=role_avg_efforts,
            supply_by_role=supply_by_role,
        )

    # ------------------------------------------------------------------
    # Project Assignments (read from Portfolio columns M-P)
    # ------------------------------------------------------------------
    # Portfolio column mapping: col_index → canonical role key
    _ASSIGNMENT_COLUMNS = {
        13: "pm",         # Column M
        14: "ba",         # Column N
        15: "functional", # Column O
        16: "technical",  # Column P
        17: "developer",  # Column Q
    }

    def read_assignments(self, active_only: bool = True) -> list[ProjectAssignment]:
        """Read person assignments from Portfolio columns M (PM), N (BA),
        O (Functional), P (Technical). Each cell contains a person name.
        Allocation defaults to 1.0 (100%) since a named person owns that
        role for the project.

        If active_only=True (default), skips completed and postponed projects
        so they don't appear in capacity/heatmap calculations.
        """
        wb = self._open()
        ws = wb["Project Portfolio"]
        assignments = []

        for row in range(4, 43):
            pid = ws.cell(row=row, column=1).value
            if not pid:
                continue

            if active_only:
                health = ws.cell(row=row, column=6).value
                pct_complete = _to_float(ws.cell(row=row, column=7).value, 0.0)
                if health and "POSTPONED" in str(health):
                    continue
                if pct_complete >= 1.0:
                    continue

            for col_idx, role_key in self._ASSIGNMENT_COLUMNS.items():
                name = ws.cell(row=row, column=col_idx).value
                if not name:
                    continue
                assignments.append(ProjectAssignment(
                    project_id=str(pid).strip(),
                    person_name=str(name).strip(),
                    role_key=role_key,
                    allocation_pct=1.0,
                ))

        return assignments

    # ------------------------------------------------------------------
    # Convenience: load everything
    # ------------------------------------------------------------------
    def load_all(self) -> dict:
        """Load all sheets and return as a dict."""
        portfolio = self.read_portfolio()
        active = [p for p in portfolio if p.is_active]
        roster = self.read_roster()
        assumptions = self.read_assumptions()
        assignments = self.read_assignments()
        return {
            "portfolio": portfolio,
            "active_portfolio": active,
            "roster": roster,
            "assumptions": assumptions,
            "assignments": assignments,
            "data_as_of": self.file_modified_time,
        }


if __name__ == "__main__":
    conn = ExcelConnector()
    data = conn.load_all()

    print(f"Data as of: {data['data_as_of']}")
    print(f"\nTotal projects: {len(data['portfolio'])}")
    print(f"Active projects: {len(data['active_portfolio'])}")
    print(f"Team members: {len(data['roster'])}")

    print("\n--- Active Projects ---")
    for p in data["active_portfolio"]:
        roles = {k: v for k, v in p.role_allocations.items() if v > 0}
        print(f"  {p.id}: {p.name} | {p.priority} | {p.est_hours}h | {p.duration_weeks}wk | roles={roles}")

    print("\n--- Team by Role ---")
    from collections import Counter
    role_counts = Counter(m.role_key for m in data["roster"])
    for role, count in sorted(role_counts.items()):
        print(f"  {role}: {count}")

    print("\n--- Supply (from RM_Assumptions) ---")
    assumptions = data["assumptions"]
    for role, supply in assumptions.supply_by_role.items():
        print(f"  {role}: {supply['headcount']} people, {supply['project_hrs_week']:.1f} proj hrs/wk")

    print(f"\n--- SDLC Phase Weights ---")
    for phase, weight in assumptions.sdlc_phase_weights.items():
        print(f"  {phase}: {weight:.0%}")

    conn.close()
