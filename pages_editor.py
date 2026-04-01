"""Project Editor page for the ETE PMO Dashboard.
Provides form-based editing of project data with validation,
writes changes back to the Excel workbook via openpyxl.
"""

from collections import defaultdict
from datetime import datetime, date
from pathlib import Path
from typing import Optional

import openpyxl
import streamlit as st

from components import (
    section_header, ROLE_DISPLAY, ROLE_ORDER,
    NAVY, BLUE, GREEN, YELLOW, RED, GRAY,
)
from data_layer import WORKBOOK_PATH
from excel_connector import PORTFOLIO_ROLE_COLUMNS, ROSTER_ROLE_MAP


# ---------------------------------------------------------------------------
# Column index map for writing back to "Project Portfolio" sheet
# ---------------------------------------------------------------------------
FIELD_COLUMNS = {
    "id": 1, "name": 2, "type": 3, "portfolio": 4, "sponsor": 5,
    "health": 6, "pct_complete": 7, "priority": 8,
    "start_date": 9, "end_date": 10, "actual_end": 11,
    "team": 12, "pm": 13, "ba": 14, "functional_lead": 15, "technical_lead": 16,
    # column 17 is skipped/unused
    "tshirt_size": 18, "est_hours": 19, "est_cost": 20,
    "alloc_ba": 21, "alloc_functional": 22, "alloc_technical": 23,
    "alloc_developer": 24, "alloc_infrastructure": 25, "alloc_dba": 26,
    "alloc_pm": 27, "alloc_wms": 28,
    "notes": 29, "sort_order": 30,
}

# Named lead → allocation role key linkage
LEAD_ALLOC_LINK = {
    "pm": {"label": "PM", "name_col": 13, "alloc_key": "alloc_pm"},
    "ba": {"label": "BA", "name_col": 14, "alloc_key": "alloc_ba"},
    "functional": {"label": "Functional Lead", "name_col": 15, "alloc_key": "alloc_functional"},
    "technical": {"label": "Technical Lead", "name_col": 16, "alloc_key": "alloc_technical"},
}

HEALTH_OPTIONS = [
    "🟢 ON TRACK", "🟡 AT RISK", "🔴 CRITICAL",
    "NEEDS TECHNICAL SPEC", "NEEDS FUNCTIONAL SPEC",
    "NOT STARTED", "COMPLETE", "POSTPONED",
]

PRIORITY_OPTIONS = ["Highest", "High", "Medium", "Low"]

TSHIRT_OPTIONS = [
    "XS: < 40 Hours", "S: 40-80 Hours", "M: 80-160 Hours",
    "L: 160-320 Hours", "XL: 320-640 Hours", "XXL: > 640 Hours",
]

TYPE_OPTIONS = ["Key Initiative", "Enhancement", "Support", "Infrastructure", "Research"]


# ---------------------------------------------------------------------------
# Excel Write-back
# ---------------------------------------------------------------------------
def _find_project_row(ws, project_id: str) -> Optional[int]:
    """Find the Excel row for a given project ID (rows 4-42)."""
    for row in range(4, 43):
        cell_val = ws.cell(row=row, column=1).value
        if cell_val and str(cell_val).strip() == project_id.strip():
            return row
    return None


def _find_empty_row(ws) -> Optional[int]:
    """Find the first empty row in the project data range (rows 4-42)."""
    for row in range(4, 43):
        if not ws.cell(row=row, column=1).value:
            return row
    return None


def _write_project(project_id: str, fields: dict, is_new: bool = False) -> Optional[str]:
    """Write project fields to the Excel workbook.
    Returns None on success, or an error message string."""
    try:
        wb = openpyxl.load_workbook(WORKBOOK_PATH)
        ws = wb["Project Portfolio"]

        if is_new:
            row = _find_empty_row(ws)
            if not row:
                wb.close()
                return "Portfolio sheet is full (max 39 projects). Cannot add new project."
        else:
            row = _find_project_row(ws, project_id)
            if not row:
                wb.close()
                return f"Could not find project {project_id} in the workbook."

        for field_key, value in fields.items():
            col = FIELD_COLUMNS.get(field_key)
            if col is None:
                continue
            # Convert dates to datetime for openpyxl
            if isinstance(value, date) and not isinstance(value, datetime):
                value = datetime.combine(value, datetime.min.time())
            ws.cell(row=row, column=col, value=value)

        wb.save(WORKBOOK_PATH)
        wb.close()
        return None

    except PermissionError:
        return "Cannot save — the workbook is open in Excel. Please close it and try again."
    except Exception as e:
        return f"Error saving: {e}"


# ---------------------------------------------------------------------------
# Health Check Warnings
# ---------------------------------------------------------------------------
def _compute_warnings(projects: list) -> list[dict]:
    """Scan all active projects for data quality issues."""
    warnings = []
    missing_alloc = defaultdict(list)

    lead_fields = [
        ("pm", "PM", "pm"),
        ("ba", "BA", "ba"),
        ("functional_lead", "Functional Lead", "functional"),
        ("technical_lead", "Technical Lead", "technical"),
    ]

    for p in projects:
        if not p.is_active:
            continue

        for attr, label, role_key in lead_fields:
            person = getattr(p, attr, None)
            alloc = p.role_allocations.get(role_key, 0.0)
            if person and alloc <= 0:
                missing_alloc[label].append(f"{p.id}: {p.name}")

        if (p.start_date or p.end_date) and p.est_hours <= 0:
            warnings.append({
                "type": "warning",
                "msg": f"{p.id}: Scheduled but missing estimated hours",
            })

        if p.est_hours > 0 and not any(v > 0 for v in p.role_allocations.values()):
            warnings.append({
                "type": "warning",
                "msg": f"{p.id}: Has {p.est_hours:.0f}h but no role allocations set",
            })

    for label, pids in missing_alloc.items():
        warnings.insert(0, {
            "type": "error",
            "msg": f"{len(pids)} project(s) have {label} assigned but 0% allocation",
            "details": pids,
        })

    return warnings


def _get_people_by_role(roster: list) -> dict:
    """Build {role_key: [name1, name2, ...]} from roster."""
    by_role = defaultdict(list)
    for m in roster:
        by_role[m.role_key].append(m.name)
    return dict(by_role)


# ---------------------------------------------------------------------------
# Render
# ---------------------------------------------------------------------------
def render(data: dict, utilization: dict, person_demand: list):
    """Render the Project Editor page."""
    all_projects = data["portfolio"]
    active = data["active_portfolio"]
    roster = data["roster"]
    people_by_role = _get_people_by_role(roster)

    # --- Portfolio Health Check ---
    warnings = _compute_warnings(all_projects)
    if warnings:
        error_count = sum(1 for w in warnings if w["type"] == "error")
        warn_count = sum(1 for w in warnings if w["type"] == "warning")

        label_parts = []
        if error_count:
            label_parts.append(f"{error_count} error{'s' if error_count > 1 else ''}")
        if warn_count:
            label_parts.append(f"{warn_count} warning{'s' if warn_count > 1 else ''}")

        with st.expander(f"⚠️ **Data Quality: {' · '.join(label_parts)}**", expanded=True):
            for w in warnings:
                if w["type"] == "error":
                    st.error(w["msg"], icon="🔴")
                else:
                    st.warning(w["msg"], icon="🟡")
                if "details" in w:
                    for detail in w["details"]:
                        st.caption(f"    → {detail}")
    else:
        st.success("All active projects pass data quality checks.", icon="✅")

    st.markdown("<div style='height: 0.5rem'></div>", unsafe_allow_html=True)

    # --- Project Selector ---
    project_options = {f"{p.id}: {p.name}": p.id for p in all_projects}
    option_labels = ["— Create New Project —"] + list(project_options.keys())

    selected_label = st.selectbox(
        "Select project to edit",
        option_labels,
        index=None,
        placeholder="Select a project to edit...",
        label_visibility="collapsed",
    )

    if selected_label is None:
        st.markdown(f"""
        <div style="text-align: center; padding: 2rem;">
            <div style="font-size: 2.5rem; margin-bottom: 0.75rem; opacity: 0.3;">✏️</div>
            <div style="font-size: 1.2rem; font-weight: 600; color: {NAVY}; margin-bottom: 0.4rem;">
                Select a Project to Edit</div>
            <div style="font-size: 0.9rem; color: {GRAY};">
                Choose a project from the dropdown above, or create a new one.</div>
        </div>
        """, unsafe_allow_html=True)
        return

    is_new = selected_label == "— Create New Project —"

    if is_new:
        project = None
    else:
        project_id = project_options[selected_label]
        project = next((p for p in all_projects if p.id == project_id), None)

    # --- Collect existing values for dropdowns ---
    existing_portfolios = sorted(set(
        p.portfolio for p in all_projects if p.portfolio
    ))
    existing_teams = sorted(set(
        p.team for p in all_projects if p.team
    ))

    # --- Form ---
    with st.form("project_editor", clear_on_submit=False):

        # Group 1: Project Identity
        section_header("Project Identity")
        id_col, name_col = st.columns(2)
        with id_col:
            proj_id = st.text_input(
                "Project ID *",
                value="" if is_new else project.id,
                disabled=not is_new,
                placeholder="e.g. ETE-125",
            )
        with name_col:
            proj_name = st.text_input(
                "Project Name *",
                value="" if is_new else project.name,
                placeholder="Enter project name",
            )

        c1, c2, c3 = st.columns(3)
        with c1:
            type_options = TYPE_OPTIONS
            type_idx = None
            if project and project.type:
                for i, t in enumerate(type_options):
                    if t == project.type:
                        type_idx = i
                        break
            proj_type = st.selectbox("Type", type_options, index=type_idx,
                                     placeholder="Select type...")
        with c2:
            priority_idx = None
            if project and project.priority:
                for i, p in enumerate(PRIORITY_OPTIONS):
                    if p == project.priority:
                        priority_idx = i
                        break
            proj_priority = st.selectbox("Priority *", PRIORITY_OPTIONS,
                                         index=priority_idx,
                                         placeholder="Select priority...")
        with c3:
            health_idx = None
            if project and project.health:
                for i, h in enumerate(HEALTH_OPTIONS):
                    if project.health.strip() in h or h in (project.health or ""):
                        health_idx = i
                        break
            proj_health = st.selectbox("Health", HEALTH_OPTIONS,
                                       index=health_idx,
                                       placeholder="Select health...")

        c4, c5 = st.columns(2)
        with c4:
            portfolio_options = existing_portfolios
            port_idx = None
            if project and project.portfolio:
                for i, po in enumerate(portfolio_options):
                    if po == project.portfolio:
                        port_idx = i
                        break
            proj_portfolio = st.selectbox("Portfolio", portfolio_options,
                                          index=port_idx,
                                          placeholder="Select portfolio...")
        with c5:
            proj_pct = st.slider(
                "% Complete",
                min_value=0, max_value=100,
                value=round(project.pct_complete * 100) if project else 0,
                format="%d%%",
            )

        # Group 2: Schedule
        section_header("Schedule & Sizing")
        sc1, sc2, sc3 = st.columns(3)
        with sc1:
            proj_start = st.date_input(
                "Start Date",
                value=project.start_date if project and project.start_date else None,
                format="MM/DD/YYYY",
            )
        with sc2:
            proj_end = st.date_input(
                "End Date",
                value=project.end_date if project and project.end_date else None,
                format="MM/DD/YYYY",
            )
        with sc3:
            tshirt_idx = None
            if project and project.tshirt_size:
                for i, t in enumerate(TSHIRT_OPTIONS):
                    if project.tshirt_size.strip() in t:
                        tshirt_idx = i
                        break
            proj_tshirt = st.selectbox("T-Shirt Size", TSHIRT_OPTIONS,
                                       index=tshirt_idx,
                                       placeholder="Select size...")

        sz1, sz2 = st.columns(2)
        with sz1:
            proj_hours = st.number_input(
                "Estimated Hours *",
                min_value=0, max_value=50000,
                value=int(project.est_hours) if project else 0,
                step=10,
            )
        with sz2:
            proj_cost = st.number_input(
                "Estimated Cost",
                min_value=0, max_value=10000000,
                value=int(project.est_cost or 0) if project else 0,
                step=100,
            )

        # Group 3: Team Assignments + Role Allocations (side-by-side)
        section_header("Team Assignments & Role Allocations")
        st.caption("When a lead is assigned, their role allocation must be > 0% for capacity planning to work.")

        lead_roles = [
            ("PM", "pm", "pm", "alloc_pm"),
            ("Business Analyst", "ba", "ba", "alloc_ba"),
            ("Functional Lead", "functional_lead", "functional", "alloc_functional"),
            ("Technical Lead", "technical_lead", "technical", "alloc_technical"),
        ]

        lead_values = {}
        alloc_values = {}

        for label, proj_attr, role_key, alloc_key in lead_roles:
            lc, rc = st.columns([3, 2])
            with lc:
                people = ["(Unassigned)"] + people_by_role.get(role_key, [])
                current = getattr(project, proj_attr, None) if project else None
                idx = 0
                if current:
                    for i, p in enumerate(people):
                        if p == current:
                            idx = i
                            break
                lead_values[proj_attr] = st.selectbox(
                    f"{label}",
                    people,
                    index=idx,
                    key=f"lead_{proj_attr}",
                )
            with rc:
                current_alloc = 0
                if project:
                    current_alloc = round(project.role_allocations.get(role_key, 0.0) * 100)
                alloc_values[alloc_key] = st.number_input(
                    f"{label} Allocation %",
                    min_value=0, max_value=100,
                    value=current_alloc,
                    step=5,
                    key=f"alloc_{role_key}",
                )

        # Additional role allocations (no named lead)
        st.markdown("**Additional Role Allocations**")
        extra_roles = [
            ("Developer", "developer", "alloc_developer"),
            ("Infrastructure", "infrastructure", "alloc_infrastructure"),
            ("DBA", "dba", "alloc_dba"),
            ("WMS Consultant", "wms", "alloc_wms"),
        ]

        er1, er2, er3, er4 = st.columns(4)
        extra_cols = [er1, er2, er3, er4]
        for i, (label, role_key, alloc_key) in enumerate(extra_roles):
            with extra_cols[i]:
                current_alloc = 0
                if project:
                    current_alloc = round(project.role_allocations.get(role_key, 0.0) * 100)
                alloc_values[alloc_key] = st.number_input(
                    f"{label} %",
                    min_value=0, max_value=100,
                    value=current_alloc,
                    step=5,
                    key=f"alloc_{role_key}",
                )

        # Group 4: Other
        section_header("Additional Details")
        o1, o2, o3 = st.columns(3)
        with o1:
            team_options = existing_teams
            team_idx = None
            if project and project.team:
                for i, t in enumerate(team_options):
                    if t == project.team:
                        team_idx = i
                        break
            proj_team = st.selectbox("Team", team_options, index=team_idx,
                                     placeholder="Select team...")
        with o2:
            proj_sponsor = st.text_input(
                "Sponsor",
                value=project.sponsor or "" if project else "",
            )
        with o3:
            proj_sort = st.number_input(
                "Sort Order",
                min_value=0, max_value=999,
                value=project.sort_order or 0 if project else 0,
            )

        proj_notes = st.text_area(
            "Notes",
            value=project.notes or "" if project else "",
            height=80,
        )

        # --- Submit ---
        submitted = st.form_submit_button("💾 Save Project", type="primary",
                                           use_container_width=True)

    # --- Validation & Save (outside form) ---
    if submitted:
        errors = []
        form_warnings = []

        # Required fields
        if not proj_id.strip():
            errors.append("Project ID is required.")
        if not proj_name.strip():
            errors.append("Project Name is required.")
        if not proj_priority:
            errors.append("Priority is required.")

        # Lead/allocation linkage validation
        lead_alloc_checks = [
            ("PM", lead_values.get("pm"), alloc_values.get("alloc_pm", 0)),
            ("BA", lead_values.get("ba"), alloc_values.get("alloc_ba", 0)),
            ("Functional Lead", lead_values.get("functional_lead"), alloc_values.get("alloc_functional", 0)),
            ("Technical Lead", lead_values.get("technical_lead"), alloc_values.get("alloc_technical", 0)),
        ]
        for label, person, alloc in lead_alloc_checks:
            if person and person != "(Unassigned)" and (alloc is None or alloc <= 0):
                errors.append(
                    f"{label} is assigned ({person}) but allocation is 0%. "
                    f"Set a percentage so capacity planning includes this role."
                )

        # Schedule validation
        if proj_start and proj_end and proj_end < proj_start:
            errors.append("End date cannot be before start date.")

        if (proj_start or proj_end) and proj_hours <= 0:
            form_warnings.append(
                "Project has dates but no estimated hours — demand will show as zero."
            )

        if proj_hours > 0 and all(v == 0 for v in alloc_values.values()):
            form_warnings.append(
                "Project has estimated hours but all role allocations are 0% — "
                "no demand will be generated."
            )

        # Show validation results
        if errors:
            for err in errors:
                st.error(err, icon="🚫")
            st.info("Please fix the errors above and save again.")
            return

        for w in form_warnings:
            st.warning(w, icon="⚠️")

        # Build the fields dict for write-back
        fields = {
            "id": proj_id.strip(),
            "name": proj_name.strip(),
            "type": proj_type,
            "portfolio": proj_portfolio,
            "sponsor": proj_sponsor.strip() if proj_sponsor else None,
            "health": proj_health,
            "pct_complete": proj_pct / 100.0,
            "priority": proj_priority,
            "start_date": proj_start if proj_start else None,
            "end_date": proj_end if proj_end else None,
            "team": proj_team,
            "tshirt_size": proj_tshirt,
            "est_hours": proj_hours,
            "notes": proj_notes.strip() if proj_notes else None,
            "sort_order": proj_sort if proj_sort else None,
        }

        # Named leads — write None for unassigned
        for proj_attr in ["pm", "ba", "functional_lead", "technical_lead"]:
            val = lead_values.get(proj_attr)
            fields[proj_attr] = val if val and val != "(Unassigned)" else None

        # Role allocations — convert from 0-100 to 0.0-1.0
        for alloc_key, pct_val in alloc_values.items():
            fields[alloc_key] = (pct_val or 0) / 100.0

        # Est cost — only write if > 0
        if proj_cost and proj_cost > 0:
            fields["est_cost"] = proj_cost

        # Write to Excel
        result = _write_project(proj_id.strip(), fields, is_new=is_new)

        if result:
            st.error(result, icon="🚫")
        else:
            st.success(
                f"{'Created' if is_new else 'Updated'} **{proj_id}: {proj_name}** successfully.",
                icon="✅",
            )
            # Clear cache so dashboard pages reflect the changes
            st.cache_data.clear()
